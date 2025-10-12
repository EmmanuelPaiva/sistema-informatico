# Edit the uploaded Excel to (1) rename header "Llave" -> "Clave" and
# (2) auto-fill missing Descripción cells using heuristics.
#
# Input file: /mnt/data/diccionario_bloques[1].xlsx
# Output file: /mnt/data/diccionario_bloques_enriquecido.xlsx

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import re

in_path = Path("/mnt/data/diccionario_bloques[1].xlsx")
out_path = Path("/mnt/data/diccionario_bloques_enriquecido.xlsx")

wb = load_workbook(in_path)
ws = wb.active  # assume single sheet named "Diccionario"

def guess_desc(field_name: str, tipo: str, clave: str, tabla: str) -> str:
    name = field_name or ""
    lower = name.lower()
    t = (tipo or "").lower()

    # Claves
    if (clave or "").upper() == "PK":
        return f"Clave primaria de la tabla {tabla}. Identifica de forma única cada registro."
    if (clave or "").upper() == "FK":
        # If it's FK but description is missing, provide a generic FK description
        return f"Clave foránea de la tabla {tabla}. Referencia a un registro relacionado."

    # IDs
    if re.search(r'(^id$|_id$|^id_|_id_|id\b)', lower):
        return f"Identificador relacionado con {tabla} o entidades asociadas."

    # Nombres / textos
    if any(k in lower for k in ["nombre","username","descripcion","detalle","texto","concepto","nota","direccion","entity","code"]):
        return f"Texto descriptivo para {name.replace('_',' ')}."

    # Correo / email / telefono
    if any(k in lower for k in ["email","correo"]):
        return "Correo electrónico de contacto."
    if "telefono" in lower or "tel" in lower:
        return "Número de teléfono de contacto."

    # Fechas y tiempos
    if "timestamp" in t or "with time zone" in t or "without time zone" in t:
        return f"Fecha y hora asociada a {name.replace('_',' ')}."
    if t == "date":
        return f"Fecha asociada a {name.replace('_',' ')}."
    if any(k in lower for k in ["created_at","updated_at","fecha","hora"]):
        return f"Marca temporal para {name.replace('_',' ')}."

    # Booleanos
    if t == "boolean" or any(k in lower for k in ["is_", "exito", "activo", "isactive", "must_change", "indicador"]):
        return f"Indicador booleano para {name.replace('_',' ')} (verdadero/falso)."

    # Númericos (monto/cantidad/total/precio/stock)
    if any(k in lower for k in ["monto","total","precio","costo","importe","presupuesto","subtotal","cantidad","stock"]):
        return f"Valor numérico para {name.replace('_',' ')}."
    if t.startswith("numeric") or t.startswith("double") or t.startswith("integer") or t.startswith("bigint") or t.startswith("smallint"):
        return f"Campo numérico para {name.replace('_',' ')}."

    # UUID
    if "uuid" in t:
        return "Identificador único universal (UUID)."

    # Por defecto
    return f"Campo {name.replace('_',' ')}."

# Iterate rows and detect table blocks
# Header row example: ["Llave","Nombre","Campo","Tipo","Tamaño","Descripción"]
# Table title row is merged in col A; we will parse "(schema)" from title.
current_table = None
current_schema = None

# First pass: rename any header "Llave" -> "Clave"
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    if row[0].value == "Llave":
        row[0].value = "Clave"

# Determine header indexes after rename
# Find the first header row to get column indexes
col_idx = {}
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    values = [c.value for c in row[:6]]
    if values and values[0] == "Clave" and values[1] == "Nombre" and values[2] == "Campo":
        for i, header in enumerate(["Clave","Nombre","Campo","Tipo","Tamaño","Descripción"], start=1):
            col_idx[header] = i
        break

if not col_idx:
    raise RuntimeError("No se encontró la fila de encabezados (Clave/Nombre/Campo/Tipo/Tamaño/Descripción).")

# Walk again and fill descriptions where empty
for r in range(1, ws.max_row + 1):
    cell_a = ws.cell(row=r, column=1)
    val_a = cell_a.value

    # Detect table title rows like "clientes (public)"
    if isinstance(val_a, str) and "(" in val_a and val_a.endswith(")") and ws.cell(row=r, column=2).value is None:
        # parse "table (schema)"
        try:
            table_part, schema_part = val_a.split("(", 1)
            current_table = table_part.strip()
            current_schema = schema_part.strip(") ").strip()
        except Exception:
            current_table = val_a.strip()
            current_schema = None
        continue

    # Skip non-data rows until we hit a header row
    if ws.cell(row=r, column=col_idx["Clave"]).value == "Clave":
        # header line, skip
        continue

    # A blank separator row between blocks?
    if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, 7)):
        continue

    # Treat as data line
    clave = ws.cell(row=r, column=col_idx["Clave"]).value
    nombre = ws.cell(row=r, column=col_idx["Nombre"]).value
    campo  = ws.cell(row=r, column=col_idx["Campo"]).value
    tipo   = ws.cell(row=r, column=col_idx["Tipo"]).value
    desc_cell = ws.cell(row=r, column=col_idx["Descripción"])

    # Only fill if empty or whitespace
    if desc_cell.value is None or (isinstance(desc_cell.value, str) and desc_cell.value.strip() == ""):
        desc_cell.value = guess_desc(campo or nombre, tipo, clave, current_table or "[tabla]")

wb.save(out_path)
out_path
