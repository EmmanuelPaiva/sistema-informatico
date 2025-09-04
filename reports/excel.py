# reports/excel.py
from datetime import datetime
from typing import Tuple
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QTableWidget, QTableView, QTreeWidget
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================================================
# =================== UTILIDADES BÁSICAS ==================
# =========================================================

def _header_style():
    fill = PatternFill("solid", fgColor="e8eef7")
    font = Font(bold=True, color="000000")
    align = Alignment(horizontal="center", vertical="center")
    bd = Side(style="thin", color="b0b0b0")
    border = Border(left=bd, right=bd, top=bd, bottom=bd)
    return {"fill": fill, "font": font, "alignment": align, "border": border}

def _data_style():
    align = Alignment(vertical="center")
    bd = Side(style="thin", color="e0e0e0")
    border = Border(left=bd, right=bd, top=bd, bottom=bd)
    return {"alignment": align, "border": border}

def autosize_columns(ws, min_width=10, max_width=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))

def _try_parse_number(val):
    """Convierte textos '1.234.567,89' o '1234,56' o '1234.56' a float si es posible."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    s_norm = s.replace(".", "").replace(",", ".")
    try:
        return float(s_norm)
    except Exception:
        return None

def _is_money_header(text):
    if not text:
        return False
    t = str(text).strip().lower()
    money_keywords = {"precio", "total", "subtotal", "importe", "monto", "costo", "unitario"}
    return any(k in t for k in money_keywords)

# =========================================================
# ============== UTILIDADES DE ESCRITURA ==================
# =========================================================

def write_kv_block(ws, start_row, title, kv_pairs):
    """
    kv_pairs: list[("Etiqueta", valor), ...]
    Escribe un bloque de 2 columnas: Etiqueta | Valor
    Devuelve la próxima fila libre (int).
    """
    ws.cell(row=start_row, column=1, value=title)
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)
    r = start_row + 2
    for k, v in kv_pairs:
        ws.cell(row=r, column=1, value=str(k)).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1
    return r + 1  # una fila en blanco

def write_table(ws, start_row, title, headers, rows, currency_cols=None, date_cols=None):
    currency_cols = set(currency_cols or [])
    date_cols = set(date_cols or [])
    # Título
    ws.cell(row=start_row, column=1, value=title)
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=13)
    # Encabezados
    hdr_fill = PatternFill("solid", fgColor="e8eef7")
    hdr_font = Font(bold=True)
    hdr_align = Alignment(horizontal="center", vertical="center")
    bd = Side(style="thin", color="b0b0b0")
    hdr_border = Border(left=bd, right=bd, top=bd, bottom=bd)

    r = start_row + 2
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = hdr_border

    # Datos
    data_align = Alignment(vertical="center")
    dat_bd = Side(style="thin", color="e0e0e0")
    dat_border = Border(left=dat_bd, right=dat_bd, top=dat_bd, bottom=dat_bd)

    r += 1
    for row in rows:
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            # moneda
            if (c - 1) in currency_cols and isinstance(val, (int, float)):
                cell.number_format = '"Gs" #,##0'
            # fecha
            if (c - 1) in date_cols and val:
                if isinstance(val, str):
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
                        try:
                            dt = datetime.strptime(val, fmt)
                            cell.value = dt
                            cell.number_format = "yyyy-mm-dd"
                            break
                        except Exception:
                            pass
                else:
                    cell.number_format = "yyyy-mm-dd"
            cell.alignment = data_align
            cell.border = dat_border
        r += 1

    return r + 1  # próxima fila libre (con una línea en blanco)

# =========================================================
# ========== EXPORT: QTableWidget / QTableView ============
# =========================================================

def export_qtable_to_excel(table_or_view, filepath, title="Exportación"):
    """
    Exporta el contenido visible de un QTableWidget o QTableView a Excel.
    - Respeta columnas ocultas.
    - Omite la columna cuyo header contenga 'Opciones'.
    - Aplica formato Gs a columnas con encabezado monetario.
    """
    if not isinstance(table_or_view, (QTableWidget, QTableView)):
        raise TypeError("Se esperaba QTableWidget o QTableView.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Título
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Modelo y conteos
    model = table_or_view.model()
    if model is None:
        wb.save(filepath)
        return

    column_count = model.columnCount()
    row_count = model.rowCount()

    # Columnas visibles
    visible_cols = []
    headers = []
    for c in range(column_count):
        if table_or_view.isColumnHidden(c):
            continue
        header_text = model.headerData(c, Qt.Horizontal, Qt.DisplayRole)
        header_text = "" if header_text is None else str(header_text)
        if header_text.strip().lower() == "opciones":
            continue
        visible_cols.append(c)
        headers.append(header_text)

    # columnas monetarias por header
    money_cols = {i for i, h in enumerate(headers) if _is_money_header(h)}

    # Escribe headers
    hdr_st = _header_style()
    start_row = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = hdr_st["fill"]
        cell.font = hdr_st["font"]
        cell.alignment = hdr_st["alignment"]
        cell.border = hdr_st["border"]

    # Escribe datos
    dat_st = _data_style()
    r = start_row + 1
    is_widget = isinstance(table_or_view, QTableWidget)

    for row in range(row_count):
        for c_idx, c in enumerate(visible_cols, start=1):
            if is_widget:
                item = table_or_view.item(row, c)
                val = item.text() if item else ""
            else:
                index = model.index(row, c)
                val = model.data(index, Qt.DisplayRole)
                if val is None:
                    val = ""
            cell = ws.cell(row=r, column=c_idx, value=val)

            # formato monetario
            if (c_idx - 1) in money_cols:
                num = _try_parse_number(val)
                if num is not None:
                    cell.value = num
                    cell.number_format = '"Gs" #,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = dat_st["alignment"]
            else:
                cell.alignment = dat_st["alignment"]
            cell.border = dat_st["border"]
        r += 1

    ws.freeze_panes = ws[f"A{start_row+1}"]
    autosize_columns(ws)
    wb.save(filepath)

# =========================================================
# =============== EXPORT: QTreeWidget =====================
# =========================================================

def export_qtree_to_excel(tree_widget: QTreeWidget, filepath, title="Exportación"):
    """
    Exporta un QTreeWidget (padres e hijos) a Excel.
    - Respeta columnas visibles.
    - Indenta la primera columna según nivel.
    - Omite columna 'Opciones'.
    - Aplica formato Gs por encabezado.
    """
    if not isinstance(tree_widget, QTreeWidget):
        raise TypeError("Se esperaba QTreeWidget.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    total_cols = tree_widget.columnCount()
    visible_cols = []
    headers = []
    header_item = tree_widget.headerItem()
    for c in range(total_cols):
        if tree_widget.isColumnHidden(c):
            continue
        text = header_item.text(c) if header_item else f"Col {c+1}"
        if text.strip().lower() == "opciones":
            continue
        visible_cols.append(c)
        headers.append(text)

    money_cols = {i for i, h in enumerate(headers) if _is_money_header(h)}

    # Headers
    hdr_st = _header_style()
    start_row = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = hdr_st["fill"]
        cell.font = hdr_st["font"]
        cell.alignment = hdr_st["alignment"]
        cell.border = hdr_st["border"]

    # Recolectar filas
    def walk(item, level=0, out=None):
        out = out or []
        row_vals = []
        for c_idx, c in enumerate(visible_cols):
            val = item.text(c)
            if c_idx == 0:
                val = ("    " * level) + (("• " if level > 0 else "") + val)
            row_vals.append(val)
        out.append(row_vals)
        for i in range(item.childCount()):
            walk(item.child(i), level + 1, out)
        return out

    data = []
    for r in range(tree_widget.topLevelItemCount()):
        data.extend(walk(tree_widget.topLevelItem(r), 0))

    # Datos
    dat_st = _data_style()
    r = start_row + 1
    for row in data:
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if (c - 1) in money_cols:
                num = _try_parse_number(val)
                if num is not None:
                    cell.value = num
                    cell.number_format = '"Gs" #,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = dat_st["alignment"]
            else:
                cell.alignment = dat_st["alignment"]
            cell.border = dat_st["border"]
        r += 1

    ws.freeze_panes = ws[f"A{start_row+1}"]
    autosize_columns(ws)
    wb.save(filepath)

# =========================================================
# ============== DETECCIÓN DE ESQUEMA (BD) ===============
# =========================================================

def _col_exists(conn, table: str, column: str) -> bool:
    """
    Devuelve True si existe la columna 'table.column' en el esquema actual.
    """
    sql = """
        SELECT 1
        FROM information_schema.columns
        WHERE table_schema = ANY (current_schemas(true))
          AND table_name = %s
          AND column_name = %s
        LIMIT 1;
    """
    with conn.cursor() as cur:
        cur.execute(sql, (table, column))
        return cur.fetchone() is not None

def _schema_hints(conn) -> Tuple[str, str]:
    """
    Determina nombres reales de columnas sensibles:
      - FK de trabajos -> obras: 'obra_id' o 'id_obra'
      - PK de gastos para ordenar: 'id_gasto' o 'id'
    """
    # trabajos -> obras
    fk_trabajos_obras = "obra_id" if _col_exists(conn, "trabajos", "obra_id") else "id_obra"
    # gastos pk
    gastos_pk = "id_gasto" if _col_exists(conn, "gastos", "id_gasto") else "id"
    return fk_trabajos_obras, gastos_pk

# =========================================================
# ============== EXPORT: OBRAS (detalle) ==================
# =========================================================

def export_obra_detalle_excel(conexion_fn, id_obra, filepath):
    """
    Exporta: Datos generales de la obra + Trabajos (resumen) + Gastos por trabajo (detalle).
    Tablas:
      - obras
      - trabajos (PK: id, FK a obras: obra_id o id_obra)
      - gastos  (FK a trabajos: trabajo_id; PK puede ser id_gasto o id)
    """
    with conexion_fn() as conn:
        fk_trabajos_obras, gastos_pk = _schema_hints(conn)

        sql_obra = """
            SELECT o.id_obra, o.nombre, o.direccion,
                   o.fecha_inicio::date, o.fecha_fin::date,
                   o.estado, o.metros_cuadrados, o.presupuesto_total, o.descripcion
            FROM obras o
            WHERE o.id_obra = %s;
        """
        # SIN porcentaje_avance
        sql_trabajos = f"""
            SELECT t.id, t.nombre, t.fecha_inicio::date, t.fecha_fin::date,
                   COALESCE(SUM(g.cantidad * g.costo_unitario), 0) AS gasto_total
            FROM trabajos t
            LEFT JOIN gastos g ON g.trabajo_id = t.id
            WHERE t.{fk_trabajos_obras} = %s
            GROUP BY t.id, t.nombre, t.fecha_inicio, t.fecha_fin
            ORDER BY t.fecha_inicio NULLS LAST, t.id;
        """
        sql_gastos_por_trabajo = f"""
            SELECT g.fecha::date, g.concepto, g.cantidad, g.unidad,
                   g.costo_unitario, (g.cantidad * g.costo_unitario) AS total
            FROM gastos g
            WHERE g.trabajo_id = %s
            ORDER BY g.fecha, g.{gastos_pk};
        """

        with conn.cursor() as cur:
            cur.execute(sql_obra, (id_obra,))
            obra = cur.fetchone()
            if not obra:
                raise ValueError(f"Obra {id_obra} no encontrada.")

            (oid, onombre, odireccion, ofinicio, ofin,
             oestado, om2, opresupuesto, odesc) = obra

            cur.execute(sql_trabajos, (id_obra,))
            trabajos = cur.fetchall()

            gastos_por_trabajo = {}
            for (id_trabajo, *_rest) in trabajos:
                cur.execute(sql_gastos_por_trabajo, (id_trabajo,))
                gastos_por_trabajo[id_trabajo] = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Obra"

    ws.cell(row=1, column=1, value=f"Obra #{oid} — {onombre}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)

    kv = [
        ("Dirección", odireccion),
        ("Estado", oestado),
        ("Fecha inicio", ofinicio),
        ("Fecha fin", ofin),
        ("Metros cuadrados", float(om2 or 0)),
        ("Presupuesto total (Gs)", float(opresupuesto or 0)),
        ("Descripción", odesc),
    ]
    r = write_kv_block(ws, start_row=3, title="Datos generales", kv_pairs=kv)

    headers_trab = ["ID Trabajo", "Nombre", "Inicio", "Fin", "Gasto total (Gs)"]
    rows_trab = []
    for (id_trabajo, nombre, finicio, ffin, gasto_total) in trabajos:
        rows_trab.append([id_trabajo, nombre, finicio, ffin, float(gasto_total or 0)])
    r = write_table(ws, start_row=r, title="Trabajos", headers=headers_trab, rows=rows_trab,
                    currency_cols={4}, date_cols={2, 3})

    for (id_trabajo, nombre, finicio, ffin, gasto_total) in trabajos:
        headers_g = ["Fecha", "Concepto", "Cantidad", "Unidad", "Costo unitario (Gs)", "Total (Gs)"]
        rows_g = []
        for (gfecha, gcon, gcant, gunid, gcosto, gtotal) in gastos_por_trabajo.get(id_trabajo, []):
            rows_g.append([gfecha, gcon, float(gcant or 0), gunid, float(gcosto or 0), float(gtotal or 0)])

        r = write_table(ws, start_row=r, title=f"Gastos — {nombre}", headers=headers_g, rows=rows_g,
                        currency_cols={4, 5}, date_cols={0})

        first_data_row = r - (len(rows_g) + 1)
        if rows_g:
            ws[f"E{r}"] = "TOTAL TRABAJO:"
            ws[f"E{r}"].font = Font(bold=True)
            ws[f"F{r}"] = f"=SUM(F{first_data_row}:F{r-1})"
            ws[f"F{r}"].number_format = '"Gs" #,##0'
            r += 2

    ws.freeze_panes = ws["A5"]
    autosize_columns(ws)
    wb.save(filepath)

# =========================================================
# ============== EXPORT: OBRAS (todas) ====================
# =========================================================

def _sheet_title(base, existing):
    """Asegura nombres de hoja <=31 chars y únicos."""
    name = (base or "Obra")[:31]
    original = name
    i = 2
    while name in existing:
        suf = f" ({i})"
        name = (original[: (31 - len(suf))] + suf)
        i += 1
    existing.add(name)
    return name

def export_todas_obras_excel(conexion_fn, filepath):
    """
    Crea un libro con 1 hoja por obra:
    - Datos generales
    - Trabajos (resumen)
    - Gastos por trabajo (detalle)
    Tablas: obras, trabajos (id, obra_id/id_obra), gastos (trabajo_id, id_gasto/id)
    """
    wb = Workbook()
    wb.remove(wb.active)
    sheet_names = set()

    with conexion_fn() as conn:
        fk_trabajos_obras, gastos_pk = _schema_hints(conn)

        SQL_OBRAS = """
            SELECT o.id_obra, o.nombre, o.direccion,
                   o.fecha_inicio::date, o.fecha_fin::date,
                   o.estado, o.metros_cuadrados, o.presupuesto_total, o.descripcion
            FROM obras o
            ORDER BY o.id_obra;
        """
        SQL_TRABAJOS = f"""
            SELECT t.id, t.nombre, t.fecha_inicio::date, t.fecha_fin::date,
                   COALESCE(SUM(g.cantidad * g.costo_unitario), 0) AS gasto_total
            FROM trabajos t
            LEFT JOIN gastos g ON g.trabajo_id = t.id
            WHERE t.{fk_trabajos_obras} = %s
            GROUP BY t.id, t.nombre, t.fecha_inicio, t.fecha_fin
            ORDER BY t.fecha_inicio NULLS LAST, t.id;
        """
        SQL_GASTOS_X_TRABAJO = f"""
            SELECT g.fecha::date, g.concepto, g.cantidad, g.unidad,
                   g.costo_unitario, (g.cantidad * g.costo_unitario) AS total
            FROM gastos g
            WHERE g.trabajo_id = %s
            ORDER BY g.fecha, g.{gastos_pk};
        """

        with conn.cursor() as cur:
            cur.execute(SQL_OBRAS)
            obras = cur.fetchall()

            for (oid, onombre, odireccion, ofinicio, ofin, oestado, om2, opresupuesto, odesc) in obras:
                ws_name = _sheet_title(f"{oid} - {onombre or 'Obra'}", sheet_names)
                ws = wb.create_sheet(ws_name)

                # Título
                ws.cell(row=1, column=1, value=f"Obra #{oid} — {onombre or ''}")
                ws.cell(row=1, column=1).font = Font(bold=True, size=16)

                # Datos generales
                kv = [
                    ("Dirección", odireccion),
                    ("Estado", oestado),
                    ("Fecha inicio", ofinicio),
                    ("Fecha fin", ofin),
                    ("Metros cuadrados", float(om2 or 0)),
                    ("Presupuesto total (Gs)", float(opresupuesto or 0)),
                    ("Descripción", odesc),
                ]
                r = write_kv_block(ws, start_row=3, title="Datos generales", kv_pairs=kv)

                # Trabajos (resumen)
                cur.execute(SQL_TRABAJOS, (oid,))
                trabajos = cur.fetchall()

                headers_trab = ["ID Trabajo", "Nombre", "Inicio", "Fin", "Gasto total (Gs)"]
                rows_trab = []
                for (id_trabajo, nombre, finicio, ffin, gasto_total) in trabajos:
                    rows_trab.append([id_trabajo, nombre, finicio, ffin, float(gasto_total or 0)])
                r = write_table(
                    ws, start_row=r, title="Trabajos",
                    headers=headers_trab, rows=rows_trab,
                    currency_cols={4}, date_cols={2, 3}
                )

                # Gastos por trabajo (detalle)
                for (id_trabajo, nombre, finicio, ffin, gasto_total) in trabajos:
                    headers_g = ["Fecha", "Concepto", "Cantidad", "Unidad", "Costo unitario (Gs)", "Total (Gs)"]
                    cur.execute(SQL_GASTOS_X_TRABAJO, (id_trabajo,))
                    gastos = cur.fetchall()
                    rows_g = []
                    for (gfecha, gcon, gcant, gunid, gcosto, gtotal) in gastos:
                        rows_g.append([gfecha, gcon, float(gcant or 0), gunid, float(gcosto or 0), float(gtotal or 0)])

                    r = write_table(
                        ws, start_row=r, title=f"Gastos — {nombre}",
                        headers=headers_g, rows=rows_g,
                        currency_cols={4, 5}, date_cols={0}
                    )

                    # Total por trabajo al final del bloque
                    if rows_g:
                        first_data_row = r - len(rows_g)
                        ws[f"E{r}"] = "TOTAL TRABAJO:"
                        ws[f"E{r}"].font = Font(bold=True)
                        ws[f"F{r}"] = f"=SUM(F{first_data_row}:F{r-1})"
                        ws[f"F{r}"].number_format = '"Gs" #,##0'
                        r += 2  # espacio entre secciones

                ws.freeze_panes = ws["A5"]
                autosize_columns(ws)

    # Guardar archivo
    wb.save(filepath)
